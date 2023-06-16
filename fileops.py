import sqlite3
import csv
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

def Setup():
    try:
        doc = open('settings.json', mode='r')
        json.load(doc)
        doc.close()
    except:
        doc = open('settings.json', mode='w')
        json.dump({}, doc)
        doc.close()
        SaveSettings("1time", "2022-11-28-07-00-00")
        SaveSettings("2time", "2022-11-28-07-00-00")
        SaveSettings("3time", "2022-11-28-07-00-00")
    try:
        DBInit()
    except:
        pass

def Reset():#DANGEROUS
    try:
        os.remove('settings.json')
    except:
        pass
    try:
        os.remove('BreadHistory.db')
    except:
        pass

def DBCursor():
    dbConnection = sqlite3.connect('BreadHistory.db')
    return dbConnection.cursor()

def DBInsert(query):
    con = sqlite3.connect('BreadHistory.db')
    cur = con.cursor()
    cur.execute(query)
    con.commit()

def DBInit():
    q1 = '''CREATE TABLE LineOne (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time DATETIME NOT NULL,
            loafs INTEGER NOT NULL,
            defective INTEGER NOT NULL)'''
    q2 = '''CREATE TABLE LineTwo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time DATETIME NOT NULL,
            loafs INTEGER NOT NULL,
            defective INTEGER NOT NULL)'''
    q3 = '''CREATE TABLE LineThree (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time DATETIME NOT NULL,
            loafs INTEGER NOT NULL,
            defective INTEGER NOT NULL)'''
    DBCursor().execute(q1)
    DBCursor().execute(q2)
    DBCursor().execute(q3)

def SaveSettings(key, value):
    doc = open('settings.json', mode='r')
    data = json.load(doc)
    doc.close()
    data.update({key: value})
    doc = open('settings.json', mode='w')
    json.dump(data, doc)
    doc.close()

def LoadSettings(key):
    doc = open('settings.json', mode='r')
    data = json.load(doc)
    doc.close
    try:
        return data[key]
    except:
        SaveSettings(key, None)
        return None

def LoadData(line):
    dirpath = LoadSettings(f"{line}path")
    nextime = LoadSettings(f"{line}time")
    while True:
        try:
            pattern = "/Detailed min</>.csv.csv"
            name = pattern.replace('</>', nextime)
            path = dirpath + name
            LoadLogFile(line, path)
        except FileNotFoundError:
            break
        else:
            log = nextime.split('-')
            nextime = datetime(int(log[0]), int(log[1]), int(log[2]), int(log[3]))
            if nextime.hour == 7:
                nextime = nextime.replace(hour=19)
            else:
                try:
                    nextime = nextime.replace(day=nextime.day + 1, hour=7)
                except:
                    try:
                        nextime = nextime.replace(month=nextime.month + 1, day=1, hour=7)
                    except:
                        nextime = nextime.replace(year=nextime.year + 1, month=1, day=1, hour=7)
            temp = []
            temp.append(str(nextime.year))
            temp.append(str(nextime.month))
            if len(temp[1]) == 1:
                temp[1] = "0" + temp[1]
            temp.append(str(nextime.day))
            if len(temp[2]) == 1:
                temp[2] = "0" + temp[2]
            temp.append(str(nextime.hour))
            if len(temp[3]) == 1:
                temp[3] = "0" + temp[3]
            temp.append("00")
            temp.append("00")
            nextime = "-".join(temp)
    log = nextime.split('-')
    nextime = datetime(int(log[0]), int(log[1]), int(log[2]), int(log[3]))
    if nextime.hour == 7:
        nextime = nextime.replace(hour=19)
    else:
        nextime = nextime.replace(hour=7)
        try:
            nextime = nextime.replace(day=nextime.day + 1)
        except:
            try:
                nextime = nextime.replace(month=nextime.month + 1, day=1)
            except:
                nextime = nextime.replace(year=nextime.year + 1, month=1, day=1)
    temp = []
    temp.append(str(nextime.year))
    temp.append(str(nextime.month))
    if len(temp[1]) == 1:
        temp[1] = "0" + temp[1]
    temp.append(str(nextime.day))
    if len(temp[2]) == 1:
        temp[2] = "0" + temp[2]
    temp.append(str(nextime.hour))
    if len(temp[3]) == 1:
        temp[3] = "0" + temp[3]
    temp.append("00")
    temp.append("00")
    nextime = "-".join(temp)
    SaveSettings(f"{line}time", nextime)

def LoadLogFile(lnum, filepath):
    if lnum == 1:
        line = 'LineOne'
    elif lnum == 2:
        line = 'LineTwo'
    elif lnum == 3:
        line = 'LineThree'
    doc = open(filepath, encoding='Windows-1251')
    reader = csv.reader(doc)
    reader.__next__()
    reader.__next__()
    for row in reader:
        row = row[0].split(';')
        log = row[0]
        dt = log.split('  ')[0].split('.')
        tm = log.split('  ')[1].split(':')
        dtm = datetime(int(dt[2]), int(dt[1]), int(dt[0]), int(tm[0]), 0)
        DBInsert(f"INSERT INTO {line} (time, loafs, defective) VALUES ('{str(dtm)}', {row[2]}, {row[1]})")

def SearchData(timestart, timeend, line, gu):
    if line == '1':
        lname = 'LineOne'
    elif line == '2':
        lname = 'LineTwo'
    elif line == '3':
        lname = 'LineThree'
    query = f"""SELECT
    strftime('%H', time) as grtime,
    time,
    SUM(loafs) as sumloafs,
    SUM(defective) as sumdefective
    FROM {lname} 
    WHERE time BETWEEN'{str(timestart)}' AND '{str(timeend)}' GROUP BY grtime"""
    return DBCursor().execute(query)

def Result(dtms, dtme, lines, unit, chart, erp, rpath, rname):
    wb = Workbook()
    wb.iso_dates = True
    ws = wb.active
    mh = [1, 'B1:C1', 'D1:E1', 'F1:G1', 'H1:0']
    for line in [1, 2, 3]:
        LoadData(line)# DB update
    if unit == 'H':
        ws['A1'] = "Время"
    else:
        ws['A1'] = "Дата"
    for l in lines:
        ws[mh[mh[0]].split(':')[0]] = f"Линия {l}"
        ws.merge_cells(mh[mh[0]])
        mh[0] += 1
    if erp[0]:
        ws[mh[mh[0]].split(':')[0]] = f"Норматив"

    setime = True
    cl = "B"
    cd = "C"
    for l in lines:
        r = "3"
        ws[cl+'2'] = "Батоны " + l
        ws[cd+'2'] = "Брак " + l
        rows = SearchData(dtms, dtme, l, unit)
        for row in rows:
            ws[cl+r] = row[2]
            ws[cd+r] = row[3]
            r = str(int(r) + 1)
            if setime:
                ws['A'+str(int(r)-1)] = row[1][:14]+'00'
        setime = False
        ws[cl+r] = f"=СУММ({cl+'2'}: {cl+str(int(r)-1)})"
        ws[cd+r] = f"=СУММ({cd+'2'}: {cd+str(int(r)-1)})"
        if cl == "B":
            cl = "D"
            cd = "E"
        elif cl == "D":
            cl = "F"
            cd = "G"
    ws['A'+r] = "Итого:"
    if chart:
        chart = LineChart()
        chart.title = ""
        chart.style = 1
        chart.y_axis.title = "Size"
        chart.y_axis.crossAx = 500
        chart.x_axis.title = "Date"
        chart.x_axis = DateAxis(crossAx=100)
        if unit == 'H':
            chart.x_axis.number_format = 'dd-mmm hh'
            chart.x_axis.majorTimeUnit = "days"
        elif unit == 'd':
            chart.x_axis.number_format = 'dd-mmm'
            chart.x_axis.majorTimeUnit = "days"
        elif unit == 'w':
            chart.x_axis.number_format = 'dd-mmm'
            chart.x_axis.majorTimeUnit = "days"
        elif unit == 'm':
            chart.x_axis.number_format = 'mmm-yyyy'
            chart.x_axis.majorTimeUnit = "months"
        data = Reference(ws, min_col=2, min_row=2, max_col=len(lines) * 2 + 1, max_row=r)
        chart.add_data(data, titles_from_data=True)
        dates = Reference(ws, min_col=1, min_row=3, max_row=r)
        chart.set_categories(dates)
        ws.add_chart(chart, 'I1')
    wb.save(rpath + "/" + rname + ".xlsx")
    return